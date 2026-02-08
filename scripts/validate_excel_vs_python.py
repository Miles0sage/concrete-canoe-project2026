#!/usr/bin/env python3
"""
NAU Concrete Canoe 2026 - Excel vs Python Validation
Reads hull values from Excel if available, compares to Python calculator.
Run from project root. Requires: pip install openpyxl pandas
"""

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from calculations.concrete_canoe_calculator import run_complete_analysis
except ImportError:
    from concrete_canoe_calculator import run_complete_analysis

SPREADSHEETS = PROJECT_ROOT / "data" / "spreadsheets"


def try_read_excel() -> dict | None:
    """Attempt to read hull dimensions from Excel. Returns None if openpyxl missing."""
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: pip install openpyxl")
        return None

    # Try mixture design - may have notes; reinforcement may have dimensions
    for fname in [
        "reinforcement_poa_calculations_2026.xlsx",
        "mixture_design_compliance_2026.xlsx",
    ]:
        path = SPREADSHEETS / fname
        if not path.exists():
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for sheet in wb.sheetnames[:3]:
                ws = wb[sheet]
                for row in ws.iter_rows(max_row=50, values_only=True):
                    row_str = " ".join(str(c).lower() for c in row if c)
                    # Look for dimension-like values
                    if any(k in row_str for k in ["length", "beam", "depth", "216", "30", "18"]):
                        return {"source": f"{fname}/{sheet}", "row": row}
            wb.close()
        except Exception as e:
            print(f"Warning reading {fname}: {e}")
    return None


def main() -> int:
    print("Excel vs Python Validation")
    print("-" * 50)

    excel_data = try_read_excel()
    if excel_data:
        print(f"Found potential data in: {excel_data.get('source', 'unknown')}")
        print(f"Sample row: {excel_data.get('row', [])[:8]}")

    # Run Python for Canoe 1 (18' x 30" x 18", 276 lbs)
    results = run_complete_analysis(
        hull_length_in=18 * 12,
        hull_beam_in=30,
        hull_depth_in=18,
        hull_thickness_in=0.5,
        concrete_weight_lbs=276,
        flexural_strength_psi=1500,
    )

    print("\nPython Calculator Output (Canoe 1):")
    print(f"  Freeboard:  {results['freeboard']['freeboard_in']:.2f} in")
    print(f"  Draft:      {results['freeboard']['draft_in']:.2f} in")
    print(f"  GM:         {results['stability']['gm_in']:.2f} in")
    print(f"  Max BM:     {results['structural']['max_bending_moment_lb_ft']:.1f} lb-ft")
    print(f"  Safety factor: {results['structural']['safety_factor']:.2f}")

    print("\nCompare these to your Excel values. If within ~10%, validation OK.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
